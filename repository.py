# Author: Tanakorn Tantanawat

"""Google Sheets persistence, with a developer-only in-memory fallback."""

from __future__ import annotations

from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


SHEETS = ("Companies", "Users", "Pages", "Questions", "QuestionOptions", "QuestionConditions", "Responses", "ResponseHistory")
STATIC_TABLES = {"Pages", "Questions", "QuestionOptions", "QuestionConditions"}
SHEET_HEADERS = {
    "Companies": ["CompanyID", "CompanyName", "Status", "LastUpdated", "SubmittedBy", "SubmittedTime"],
    "Users": ["Email", "Name", "CompanyID"],
    "Pages": ["PageID", "PageTitle"],
    "Questions": ["QuestionID", "PageID", "Sequence", "QuestionText", "AnswerType", "Required", "Active"],
    "QuestionOptions": ["OptionID", "QuestionID", "DisplayOrder", "OptionValue", "DisplayText"],
    "QuestionConditions": ["LogicID", "QuestionID", "Seq", "LeftParen", "DependsOnQuestion", "Operator", "ExpectedValue", "RightParen", "LogicalWithNext"],
    "Responses": ["CompanyID", "QuestionID", "ResponseValue", "LastModifiedBy", "LastModifiedTime"],
    "ResponseHistory": ["HistoryID", "CompanyID", "QuestionID", "OldValue", "NewValue", "ModifiedBy", "ModifiedTime"],
}


def default_workbook_path() -> Path:
    return Path(__file__).resolve().parent.parent / "db" / "dlml_assessment_db.xlsx"


def now() -> str:
    return datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M:%S %z")


SEED: dict[str, list[dict[str, str]]] = {
    "Companies": [
        {"CompanyID": "C001", "CompanyName": "Northwind Pumps", "Status": "Draft", "LastUpdated": "", "SubmittedBy": "", "SubmittedTime": ""},
        {"CompanyID": "C002", "CompanyName": "Contoso Manufacturing", "Status": "Draft", "LastUpdated": "", "SubmittedBy": "", "SubmittedTime": ""},
    ],
    "Users": [
        {"Email": "maya@northwind.example", "Name": "Maya Chen", "CompanyID": "C001"},
        {"Email": "noah@northwind.example", "Name": "Noah Patel", "CompanyID": "C001"},
        {"Email": "sofia@contoso.example", "Name": "Sofia Reyes", "CompanyID": "C002"},
    ],
    "Pages": [
        {"PageID": "P001", "PageTitle": "Business Goal"},
        {"PageID": "P002", "PageTitle": "Business Objectives"},
    ],
    "Questions": [
        {"QuestionID": "Q001", "PageID": "P001", "Sequence": "1", "QuestionText": "Describe your company", "AnswerType": "Text", "Required": "TRUE", "Active": "TRUE"},
        {"QuestionID": "Q002", "PageID": "P002", "Sequence": "2", "QuestionText": "Lean implementation level", "AnswerType": "Choice", "Required": "TRUE", "Active": "TRUE"},
        {"QuestionID": "Q003", "PageID": "P002", "Sequence": "3", "QuestionText": "Is TPM implemented?", "AnswerType": "Choice", "Required": "TRUE", "Active": "TRUE"},
    ],
    "QuestionOptions": [
        {"OptionID": "O002A", "QuestionID": "Q002", "DisplayOrder": "1", "OptionValue": "NONE", "DisplayText": "None"},
        {"OptionID": "O002B", "QuestionID": "Q002", "DisplayOrder": "2", "OptionValue": "SOME", "DisplayText": "Some"},
        {"OptionID": "O002C", "QuestionID": "Q002", "DisplayOrder": "3", "OptionValue": "ALL", "DisplayText": "All"},
        {"OptionID": "O003A", "QuestionID": "Q003", "DisplayOrder": "1", "OptionValue": "NO", "DisplayText": "No"},
        {"OptionID": "O003B", "QuestionID": "Q003", "DisplayOrder": "2", "OptionValue": "YES", "DisplayText": "Yes"},
    ],
    "QuestionConditions": [
        {"LogicID": "L001", "QuestionID": "Q003", "Seq": "1", "LeftParen": "", "DependsOnQuestion": "Q002", "Operator": "=", "ExpectedValue": "ALL", "RightParen": "", "LogicalWithNext": "END"}
    ],
    "Responses": [],
    "ResponseHistory": [],
}


class ExcelRepository:
    """Repository backed by the local Excel workbook for local development."""

    def __init__(self, workbook_path: str | Path | None = None) -> None:
        self.workbook_path = Path(workbook_path or default_workbook_path()).resolve()
        self._workbook = self._load_workbook()
        self._rows_cache: dict[str, list[dict[str, str]]] = {}
        self._design_cache: dict[str, Any] | None = None
        self._login_cache: dict[str, list[dict[str, str]]] = {}
        self._history_counter = 0

    def _load_workbook(self) -> Any:
        if self.workbook_path.exists():
            workbook = load_workbook(self.workbook_path, data_only=False)
        else:
            workbook = Workbook()
            workbook.remove(workbook.active)
            for sheet_name in SHEETS:
                worksheet = workbook.create_sheet(title=sheet_name)
                headers = SHEET_HEADERS.get(sheet_name, [])
                if headers:
                    worksheet.append(headers)
                for row in SEED.get(sheet_name, []):
                    worksheet.append([row.get(header, "") for header in headers])
            self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(self.workbook_path)

        for sheet_name in SHEETS:
            if sheet_name not in workbook.sheetnames:
                workbook.create_sheet(title=sheet_name)
            worksheet = workbook[sheet_name]
            if worksheet.max_row == 0:
                headers = SHEET_HEADERS.get(sheet_name, [])
                if headers:
                    worksheet.append(headers)
        return workbook

    def _sheet_headers(self, worksheet: str) -> list[str]:
        ws = self._workbook[worksheet]
        if ws.max_row == 0:
            return []
        return ["" if cell is None else str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    def _rows_for_sheet(self, worksheet: str) -> list[dict[str, str]]:
        ws = self._workbook[worksheet]
        headers = self._sheet_headers(worksheet)
        if not headers:
            return []
        rows: list[dict[str, str]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            row_values = ["" if value is None else str(value) for value in row]
            rows.append({headers[index]: row_values[index] if index < len(row_values) else "" for index in range(len(headers))})
        return rows

    def rows(self, worksheet: str) -> list[dict[str, str]]:
        if worksheet in STATIC_TABLES:
            if worksheet not in self._rows_cache:
                self._rows_cache[worksheet] = deepcopy(self._rows_for_sheet(worksheet))
            return deepcopy(self._rows_cache[worksheet])
        return deepcopy(self._rows_for_sheet(worksheet))

    def clear_cache(self) -> None:
        self._rows_cache.clear()
        self._design_cache = None
        self._login_cache.clear()
        self._history_counter = 0

    def load_login_data(self, force: bool = False) -> None:
        if self._login_cache and not force:
            return
        self._login_cache = {
            "Companies": self.rows("Companies"),
            "Users": self.rows("Users"),
        }

    def design_data(self) -> dict[str, Any]:
        if self._design_cache is None:
            questions = self.rows("Questions")
            options_by_question: dict[str, list[dict[str, str]]] = {}
            conditions_by_question: dict[str, list[dict[str, str]]] = {}
            for row in self.rows("QuestionOptions"):
                options_by_question.setdefault(row["QuestionID"], []).append(row)
            for row in self.rows("QuestionConditions"):
                conditions_by_question.setdefault(row["QuestionID"], []).append(row)
            try:
                pages = self.rows("Pages")
            except KeyError:
                pages = []
            self._design_cache = {
                "Pages": pages,
                "Questions": questions,
                "OptionsByQuestion": options_by_question,
                "ConditionsByQuestion": conditions_by_question,
            }
        return self._design_cache

    def _find_row(self, worksheet: str, predicate) -> tuple[int | None, dict[str, str] | None]:
        ws = self._workbook[worksheet]
        headers = self._sheet_headers(worksheet)
        if not headers:
            return None, None
        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            row_values = ["" if value is None else str(value) for value in row]
            payload = {headers[index]: row_values[index] if index < len(row_values) else "" for index in range(len(headers))}
            if predicate(payload):
                return row_number, payload
        return None, None

    def _write_value(self, worksheet: str, row_number: int, header: str, value: str) -> None:
        ws = self._workbook[worksheet]
        headers = self._sheet_headers(worksheet)
        if header not in headers:
            return
        column_index = headers.index(header) + 1
        ws.cell(row=row_number, column=column_index, value=value)

    def company(self, company_id: str) -> dict[str, str] | None:
        self.load_login_data()
        return next((row for row in self._login_cache["Companies"] if row["CompanyID"] == company_id), None)

    def company_by_name(self, company_name: str) -> dict[str, str] | None:
        self.load_login_data()
        target = company_name.strip().casefold()
        return next((row for row in self._login_cache["Companies"] if row["CompanyName"].strip().casefold() == target), None)

    def user(self, email: str, company_id: str) -> dict[str, str] | None:
        self.load_login_data()
        return next((row for row in self._login_cache["Users"] if row["Email"].lower() == email.lower() and row["CompanyID"] == company_id), None)

    def responses_for(self, company_id: str) -> dict[str, str]:
        return {row["QuestionID"]: row["ResponseValue"] for row in self.rows("Responses") if row["CompanyID"] == company_id}

    def save_response(self, company_id: str, question_id: str, value: str, email: str) -> bool:
        row_number, current = self._find_row("Responses", lambda row: row.get("CompanyID") == company_id and row.get("QuestionID") == question_id)
        old = current.get("ResponseValue", "") if current else ""
        if old == value:
            return False
        stamp = now()
        sheet = self._workbook["Responses"]
        headers = self._sheet_headers("Responses")
        if not headers:
            headers = SHEET_HEADERS["Responses"]
            sheet.append(headers)
        if current and row_number is not None:
            self._write_value("Responses", row_number, "ResponseValue", value)
            self._write_value("Responses", row_number, "LastModifiedBy", email)
            self._write_value("Responses", row_number, "LastModifiedTime", stamp)
        else:
            sheet.append([company_id, question_id, value, email, stamp])
        self._update_company_last_updated(company_id, stamp)
        self.append_response_history(company_id, question_id, old, value, email, stamp)
        self._workbook.save(self.workbook_path)
        self.clear_cache()
        return True

    def _update_company_last_updated(self, company_id: str, stamp: str) -> None:
        row_number, current = self._find_row("Companies", lambda row: row.get("CompanyID") == company_id)
        if not current or row_number is None:
            return
        self._write_value("Companies", row_number, "LastUpdated", stamp)

    def append_response_history(self, company_id: str, question_id: str, old_value: str, new_value: str, email: str, stamp: str) -> None:
        self._history_counter += 1
        history = self._workbook["ResponseHistory"]
        history.append([f"H{self._history_counter:04}", company_id, question_id, old_value, new_value, email, stamp])

    def submit(self, company_id: str, email: str) -> None:
        row_number, current = self._find_row("Companies", lambda row: row.get("CompanyID") == company_id)
        if not current or row_number is None:
            raise ValueError("Company does not exist")
        stamp = now()
        self._write_value("Companies", row_number, "Status", "Submitted")
        self._write_value("Companies", row_number, "LastUpdated", stamp)
        self._write_value("Companies", row_number, "SubmittedBy", email)
        self._write_value("Companies", row_number, "SubmittedTime", stamp)
        self._workbook.save(self.workbook_path)
        self.clear_cache()


class InMemoryRepository:
    """For local UI testing only. Each Streamlit session receives its own copy."""

    def __init__(self) -> None:
        self.tables = deepcopy(SEED)
        self._static_cache: dict[str, list[dict[str, str]]] = {}
        self._design_cache: dict[str, Any] | None = None
        self._login_cache: dict[str, list[dict[str, str]]] = {}
        self._history_counter = 0

    def rows(self, worksheet: str) -> list[dict[str, str]]:
        if worksheet in STATIC_TABLES:
            if worksheet not in self._static_cache:
                self._static_cache[worksheet] = deepcopy(self.tables[worksheet])
            return deepcopy(self._static_cache[worksheet])
        return self.tables[worksheet]

    def clear_cache(self) -> None:
        self._static_cache.clear()
        self._design_cache = None
        self._login_cache.clear()
        self._history_counter = 0

    def load_login_data(self, force: bool = False) -> None:
        if self._login_cache and not force:
            return
        self._login_cache = {
            "Companies": self.rows("Companies"),
            "Users": self.rows("Users"),
        }

    def design_data(self) -> dict[str, Any]:
        if self._design_cache is None:
            questions = self.rows("Questions")
            options_by_question: dict[str, list[dict[str, str]]] = {}
            conditions_by_question: dict[str, list[dict[str, str]]] = {}
            for row in self.rows("QuestionOptions"):
                options_by_question.setdefault(row["QuestionID"], []).append(row)
            for row in self.rows("QuestionConditions"):
                conditions_by_question.setdefault(row["QuestionID"], []).append(row)
            try:
                pages = self.rows("Pages")
            except KeyError:
                pages = []
            self._design_cache = {
                "Pages": pages,
                "Questions": questions,
                "OptionsByQuestion": options_by_question,
                "ConditionsByQuestion": conditions_by_question,
            }
        return self._design_cache

    def company(self, company_id: str) -> dict[str, str] | None:
        self.load_login_data()
        return next((row for row in self._login_cache["Companies"] if row["CompanyID"] == company_id), None)

    def company_by_name(self, company_name: str) -> dict[str, str] | None:
        self.load_login_data()
        target = company_name.strip().casefold()
        return next((row for row in self._login_cache["Companies"] if row["CompanyName"].strip().casefold() == target), None)

    def user(self, email: str, company_id: str) -> dict[str, str] | None:
        self.load_login_data()
        return next((row for row in self._login_cache["Users"] if row["Email"].lower() == email.lower() and row["CompanyID"] == company_id), None)

    def responses_for(self, company_id: str) -> dict[str, str]:
        return {row["QuestionID"]: row["ResponseValue"] for row in self.rows("Responses") if row["CompanyID"] == company_id}

    def save_response(self, company_id: str, question_id: str, value: str, email: str) -> bool:
        existing = next((row for row in self.rows("Responses") if row["CompanyID"] == company_id and row["QuestionID"] == question_id), None)
        old = existing["ResponseValue"] if existing else ""
        if old == value:
            return False
        stamp = now()
        if existing:
            existing.update({"ResponseValue": value, "LastModifiedBy": email, "LastModifiedTime": stamp})
        else:
            self.rows("Responses").append({"CompanyID": company_id, "QuestionID": question_id, "ResponseValue": value, "LastModifiedBy": email, "LastModifiedTime": stamp})
        self.append_response_history(company_id, question_id, old, value, email, stamp)
        return True

    def append_response_history(self, company_id: str, question_id: str, old_value: str, new_value: str, email: str, stamp: str) -> None:
        self._history_counter += 1
        self.rows("ResponseHistory").append({"HistoryID": f"H{self._history_counter:04}", "CompanyID": company_id, "QuestionID": question_id, "OldValue": old_value, "NewValue": new_value, "ModifiedBy": email, "ModifiedTime": stamp})

    def submit(self, company_id: str, email: str) -> None:
        company = self.company(company_id)
        assert company
        stamp = now()
        company.update({"Status": "Submitted", "LastUpdated": stamp, "SubmittedBy": email, "SubmittedTime": stamp})


class GoogleSheetsRepository:
    """Repository for the production Google Spreadsheet contract."""

    def __init__(self, spreadsheet: Any) -> None:
        self.spreadsheet = spreadsheet
        self._rows_cache: dict[str, list[dict[str, str]]] = {}
        self._design_cache: dict[str, Any] | None = None
        self._login_cache: dict[str, list[dict[str, str]]] = {}
        self._history_counter = 0

    @classmethod
    def from_service_account(cls, service_account_info: dict[str, Any], spreadsheet_id: str) -> "GoogleSheetsRepository":
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        credentials = Credentials.from_service_account_info(service_account_info, scopes=scopes)
        return cls(gspread.authorize(credentials).open_by_key(spreadsheet_id))

    def rows(self, worksheet: str) -> list[dict[str, str]]:
        if worksheet not in self._rows_cache:
            self._rows_cache[worksheet] = self.spreadsheet.worksheet(worksheet).get_all_records(default_blank="")
        return deepcopy(self._rows_cache[worksheet])

    def clear_cache(self) -> None:
        self._rows_cache.clear()
        self._design_cache = None
        self._login_cache.clear()
        self._history_counter = 0

    def load_login_data(self, force: bool = False) -> None:
        if force:
            self._rows_cache.pop("Companies", None)
            self._rows_cache.pop("Users", None)
            self._login_cache.clear()
        if self._login_cache and not force:
            return
        self._login_cache = {
            "Companies": self.rows("Companies"),
            "Users": self.rows("Users"),
        }

    def _worksheet(self, name: str):
        return self.spreadsheet.worksheet(name)

    def design_data(self) -> dict[str, Any]:
        if self._design_cache is None:
            questions = self.rows("Questions")
            options_by_question: dict[str, list[dict[str, str]]] = {}
            conditions_by_question: dict[str, list[dict[str, str]]] = {}
            for row in self.rows("QuestionOptions"):
                options_by_question.setdefault(row["QuestionID"], []).append(row)
            for row in self.rows("QuestionConditions"):
                conditions_by_question.setdefault(row["QuestionID"], []).append(row)
            try:
                pages = self.rows("Pages")
            except KeyError:
                pages = []
            self._design_cache = {
                "Pages": pages,
                "Questions": questions,
                "OptionsByQuestion": options_by_question,
                "ConditionsByQuestion": conditions_by_question,
            }
        return self._design_cache

    def _find_row(self, worksheet: str, predicate) -> tuple[int, dict[str, str]] | tuple[None, None]:
        for row_number, row in enumerate(self.rows(worksheet), start=2):
            if predicate(row):
                return row_number, row
        return None, None

    def company(self, company_id: str) -> dict[str, str] | None:
        self.load_login_data()
        return next((row for row in self._login_cache["Companies"] if row["CompanyID"] == company_id), None)

    def company_by_name(self, company_name: str) -> dict[str, str] | None:
        self.load_login_data()
        target = company_name.strip().casefold()
        return next((row for row in self._login_cache["Companies"] if row["CompanyName"].strip().casefold() == target), None)

    def user(self, email: str, company_id: str) -> dict[str, str] | None:
        self.load_login_data()
        return next((row for row in self._login_cache["Users"] if row["Email"].lower() == email.lower() and row["CompanyID"] == company_id), None)

    def responses_for(self, company_id: str) -> dict[str, str]:
        return {row["QuestionID"]: row["ResponseValue"] for row in self.rows("Responses") if row["CompanyID"] == company_id}

    def save_response(self, company_id: str, question_id: str, value: str, email: str) -> bool:
        row_number, current = self._find_row("Responses", lambda row: row["CompanyID"] == company_id and row["QuestionID"] == question_id)
        old = current["ResponseValue"] if current else ""
        if old == value:
            return False
        stamp = now()
        sheet = self._worksheet("Responses")
        if current:
            headers = sheet.row_values(1)
            updated = {**current, "ResponseValue": value, "LastModifiedBy": email, "LastModifiedTime": stamp}
            sheet.update(f"A{row_number}", [[updated.get(header, "") for header in headers]])
        else:
            sheet.append_row([company_id, question_id, value, email, stamp], value_input_option="USER_ENTERED")
        self._update_company_last_updated(company_id, stamp)
        self.append_response_history(company_id, question_id, old, value, email, stamp)
        self._rows_cache.pop("Responses", None)
        return True

    def _update_company_last_updated(self, company_id: str, stamp: str) -> None:
        company_row_number, current = self._find_row("Companies", lambda row: row["CompanyID"] == company_id)
        if not current:
            return
        sheet = self._worksheet("Companies")
        headers = sheet.row_values(1)
        updated = {**current, "LastUpdated": stamp}
        sheet.update(f"A{company_row_number}", [[updated.get(header, "") for header in headers]])
        self._rows_cache.pop("Companies", None)
        if "Companies" in self._login_cache:
            for row in self._login_cache["Companies"]:
                if row["CompanyID"] == company_id:
                    row.update({"LastUpdated": stamp})
                    break

    def append_response_history(self, company_id: str, question_id: str, old_value: str, new_value: str, email: str, stamp: str) -> None:
        self._history_counter += 1
        history = self._worksheet("ResponseHistory")
        history.append_row([f"H{self._history_counter:04}", company_id, question_id, old_value, new_value, email, stamp], value_input_option="USER_ENTERED")

    def submit(self, company_id: str, email: str) -> None:
        row_number, current = self._find_row("Companies", lambda row: row["CompanyID"] == company_id)
        if not current:
            raise ValueError("Company does not exist")
        stamp = now()
        sheet = self._worksheet("Companies")
        headers = sheet.row_values(1)
        updated = {**current, "Status": "Submitted", "LastUpdated": stamp, "SubmittedBy": email, "SubmittedTime": stamp}
        sheet.update(f"A{row_number}", [[updated.get(header, "") for header in headers]])
        self._rows_cache.pop("Companies", None)
        if "Companies" in self._login_cache:
            for row in self._login_cache["Companies"]:
                if row["CompanyID"] == company_id:
                    row.update({"Status": "Submitted", "LastUpdated": stamp, "SubmittedBy": email, "SubmittedTime": stamp})
                    break
